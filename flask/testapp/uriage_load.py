import pathlib
import openpyxl
from testapp.models.uriage import Uriage
from testapp import session
import datetime
import os
import re
from sqlalchemy import distinct

class UriageLoad:
    
    def chgdate(self, val):
        if isinstance(val, datetime.datetime):
            return val.strftime('%Y/%m/%d')
        return ''
        
    def chkint(self, val):
        if type(val) is int:
            return val
        return None

    def chkstr(self, val):
        if type(val) is str:
            return val;
        return ""

    def convstr(self, val, keta=None):
        if type(val) is int:
            if keta is not None:
                return str(val).zfill(4)
            return str(val)
        elif type(val) is None:
            return ""
        return val

    def load(self, files):

        upload_error = self.check(files)
        if upload_error != '':
            return upload_error

        for file in files:
            path = os.path.join('files', file)
            path_obj = pathlib.Path(path)
            # workbook obj
            wb = openpyxl.load_workbook(path_obj, data_only=True)
            # [roop] worksheet obj
            for sh in wb:
                for rownum in range(9, 19):
                    if sh.cell(rownum, 2).value is None:
                        continue

                    uriage = Uriage()
                    uriage.denpyo_no = self.convstr(sh.cell(2,7).value)
                    uriage.date = self.chgdate(sh.cell(3,7).value)
                    uriage.torihikisaki_code = self.convstr(sh.cell(4, 3).value, keta=4)
                    uriage.torihikisaki_mei = re.sub('\s*御中', '', self.chkstr(sh.cell(3, 2).value))
                    uriage.tantosha_code = self.convstr(sh.cell(7, 8).value)
                    uriage.tantosha_mei = self.chkstr(sh.cell(7, 7).value)
                    uriage.no = self.chkint(sh.cell(rownum, 1).value)
                    uriage.shohin_code = self.convstr(sh.cell(rownum, 2).value)
                    uriage.shohin_mei = self.chkstr(sh.cell(rownum, 3).value)
                    uriage.suryo = self.chkint(sh.cell(rownum, 4).value)
                    uriage.tanka = self.chkint(sh.cell(rownum, 5).value)
                    if uriage.suryo is not None and uriage.tanka is not None:
                        uriage.kingaku = uriage.suryo * uriage.tanka
                    uriage.biko = self.chkstr(sh.cell(rownum, 7).value)
                    uriage.filename = file
                    session.add(uriage)

        session.commit()
        return ''

    def check(self, files):

        for file in files:
            path = os.path.join('files', file)
            path_obj = pathlib.Path(path)
            # workbook obj
            wb = openpyxl.load_workbook(path_obj, data_only=True)
            denpyono_list = []
            # [roop] worksheet obj
            for sh in wb:

                uriage = Uriage()
                title = self.chkstr(sh.cell(1, 1).value)
                if title is None or title != '売上伝票':
                    return 'アップロードファイルが売上伝票ではありません'
                denpyono_list.append(self.convstr(sh.cell(2,7).value))

        results = session.query(distinct(Uriage.denpyo_no)).filter(Uriage.denpyo_no.in_(denpyono_list)).all()
        if len(results) > 0:
            return '同一伝票番号の売上伝票がすでに登録されています'

        return ''
